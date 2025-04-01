import openpyxl
import os
from datetime import datetime

class KakeiboWriter:
    def __init__(self, date):
        #folder_path = r'C:\Users\tawab\OneDrive\夫婦共有\書類\家計簿'
        folder_path = os.path.join(os.getenv('USERPROFILE'), "OneDrive", "夫婦共有", "書類", "家計簿")
        self.file_path = folder_path + "\家計簿_" + date.split('/')[0] + ".xlsx"
        self.wb = openpyxl.load_workbook(self.file_path)
        self.ws = self.wb["費目"]

        self.column_mapping = {
            'B': '購入日',
            'C': '金額',
            'D': '費目',
            'E': '名称',
            'F': '非クレカ使用',
            'G': '立替者'
        }

    def get_actual_last_row(self):
        for row in range(7, self.ws.max_row + 1):  # データは7行目～
            if self.ws[f'B{row}'].value is None:
                return row - 1
        return self.ws.max_row

    def append_data(self, form):
        last_row = self.get_actual_last_row()
        new_row = last_row + 1

        for col, key in self.column_mapping.items():
            value = form.get(key, '')

            # 購入日が日付の場合に日付型に変換して設定
            if key == '購入日' and value:
                value = datetime.strptime(value, "%Y/%m/%d").date()  # 日付型に変換

            self.ws[f'{col}{new_row}'] = value

        self.wb.save(self.file_path)
        return new_row

if __name__ == "__main__":
    
    form = {
        '購入日': '2025/3/20',
        '費目': '食費',
        '金額': 100,
        '名称': '仮入力',
        '非クレカ使用': '✓',
        '立替者': ''
    }

    writer = KakeiboWriter(form['購入日'])
    new_row = writer.append_data(form)
    print(new_row)
